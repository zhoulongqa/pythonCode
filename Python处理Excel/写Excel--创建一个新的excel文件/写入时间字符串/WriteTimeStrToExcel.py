#encoding=utf-8
import openpyxl
import time
out=openpyxl.Workbook()
sheet2=out.create_sheet(u"员工工资表",index=1)
print u"创建表的表名：",sheet2.title

now=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
sheet2.cell(coordinate="C2").value=now
print sheet2.cell(coordinate="C2").value
out.save("D:\\python\\gloryroadtext.xlsx")