#encoding=utf-8
import openpyxl
import time
out=openpyxl.Workbook()
sheet2=out.create_sheet(u"员工工资表",index=1)
print u"创建表的表名：",sheet2.title

#指定写列excel
for i in xrange(1,5):
    sheet2.cell(coordinate="A%s" %(i)).value=str(i)+'kk'
    print sheet2.cell(coordinate="A%s" %(i)).value
out.save("D:\\python\\gloryroadtext.xlsx")