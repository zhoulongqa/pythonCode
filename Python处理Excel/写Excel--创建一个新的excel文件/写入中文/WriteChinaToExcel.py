#encoding=utf-8
'''
针对写入中文，需要注意的是，excel中默认中文
'''
import openpyxl
out=openpyxl.Workbook()
sheet2=out.create_sheet(u"员工工资表",index=1)
print u"创建表的表名：",sheet2.title
#写入中文字符
sheet2.cell(coordinate="C2").value="这里是中文"
print sheet2.cell(coordinate="C2").value
out.save("D:\\python\\gloryroadtext.xlsx")
