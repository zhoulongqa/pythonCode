#encoding=utf-8
import openpyxl
out=openpyxl.Workbook()
sheet2=out.create_sheet(u"员工工资表",index=1)
print type(sheet2)
print u"创建表的表名：",sheet2.title
sheet2.cell(row=1,column=2).value=1
print sheet2.cell(row=1,column=2).value
'''
sheet2.cell(coordinate='C3').value=97978
print sheet2.cell(coordinate='C3').value
'''
out.save("D;\\python\\gloryroadtext.xlsx")