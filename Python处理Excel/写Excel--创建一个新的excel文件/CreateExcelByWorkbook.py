#encoding=utf-8
'''
创建新的工作表
out.create_sheet(title=None,index=None)
使用该方法来创建一个空的工作表
参数说明：
title：表名，是Unicode字符串
index：可选参数，表示工作表的索引号，整型
'''
import openpyxl
out=openpyxl.Workbook()
sheet=out.create_sheet(u"员工工资表",index=1)
print type(sheet)
print u"创建表的表名：",sheet.title

