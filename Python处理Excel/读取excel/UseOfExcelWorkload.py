#encoding=utf-8
'''
读取excel
workbook=load_workbook(filename)
该方法是将要读取的excel文件加载到内存，并返回文档对象workbook
filename：表示要解析的excel名，可以是绝对路径也可以是相对路径，创建test.xlsx文件
'''
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding("utf-8")
workbook=load_workbook(r"D:\\python\\test.xlsx")
print workbook

print "---------------获取excel中所有sheet的名字---------------"
'''
workbook.get_sheet_names()
该方法获取excel文件中所有工作表的名字组成的对象，返回的是一个list
'''
#获取excel中所有工作表的名字
bookNameList=workbook.get_sheet_names()
print type(bookNameList)
print bookNameList
for i in bookNameList:
    print i,
print "\n"
print "=====================获取一个工作表(sheet)==========================="
'''
workbook.get_sheet_by_name(name)
该方法通过表名获取该工作表的对象sheet
参数name表示工作表的名字，如果为中文，需要在前面加上字符"u",指定为Unicode字符
'''
#通过表名获取指定的工作表
sheet1=workbook.get_sheet_by_name(u"员工信息表")
print type(sheet1)

print "***********************获取单元格中的所有行********************"
'''
openpyxl会根据工作表中实际有数据的区域来确定行数和列数
sheet.rows
该属性表示获取工作表中所有有效的行，返回的是一个tuple，一行就是tuple中的一个元素，这些元素也是一个
tuple，里面存着这一行中所有单元格的对象
'''
#获取工作表中所有的行
all_row=sheet1.rows
print type(all_row)
#print len(all_row)
for i in all_row:
    print "\n",type(i)
    for j in i:
        print j,    #单元格对象
    #print j.value   #打印单元格的内容


'''
sheet.columns
该属性表示获取工作表中所有有效的列，返回的是一个tuple，一列就是tuple中的一个
元素，这些元素也是一个tuple，里面存着这一列中所有单元格的对象
'''
print "-------------------------获取单元格中所有的列-------------"
print "*************************获取单元格中所有的列**************"
#获取单元格中所有的列
all_col=sheet1.columns
print type(all_col)
for i in all_col:
    print "\n",type(i)
    for j in i:
        print j

print "=====================获取某个单元格=========================="
print "----------------通过行获取单元格-----------------"
#获取工作表中所有的行
all_row=sheet1.rows
#获取单元格中坐标为C3的单元格
for i in all_row:
    for j in xrange(len(i)):
        if i[j].coordinate=="C3":
            print u"单元格中C3的值：",i[j].value
            break
print "----------------通过列获取单元格----------------"
#获取党员个中所有的列
all_col=sheet1.columns
#获取单元格 中坐标为A3的单元格
for i in all_col:
    for j in xrange(len(i)):
        #print i[j].value
        if i[j].coordinate=="A3":
            print u"单元格中A3的值：",i[j].value
            break
print u"查看列表中的全部内容"
all_row=sheet1.rows
for i in all_row:
    for j in xrange(len(i)):
        print str(i[j].value).center(15),
    print ""

print "---------------通过坐标获取某个单元格------------------"
'''
通过坐标获取单元格使用的方法为：sheet.cell()
函数原型：
cell(self,coordinate=None,row=None,column=None,value=None)
该方法，通过横纵坐标获取单元格对象
参数说明：
coordinate:单元格的位置参数，包括横纵坐标，比如A1
row：单元格的横坐标，整数，从1开始计算，表示第一行
column：单元格的纵坐标，整数，从1开始计算，表示第一列
获取某个单元格的对象又两种方法：
1、直接设置coordinate来获得某个单元格的对象
2、设置row和column两个参数来获取单元格的对象
如果同时设置了coordinate和row、column，将会返回coordinate参数所在单元格的对象
'''
#获取单元格B3
b3=sheet1.cell(coordinate="B3")
print u"单元格B3的值：",b3.value

#获取单元格C2
c2=sheet1.cell(row=2,column=3)
print u"单元格C2的值：",c2.value

#同时指定coordinate，row和column
c=sheet1.cell(coordinate="C3",row=3,column=3)
print c.value

print "****************通过行列获取一片区域****************"
'''
rows=sheet1.rows
#获取一片区域
row_two=rows[:2]
print row_two
'''



