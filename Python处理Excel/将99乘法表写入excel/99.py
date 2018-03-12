#encoding=utf-8
import openpyxl
out=openpyxl.Workbook()
sheet=out.create_sheet(u"99乘法表",index=1)
print type(sheet)
print u"创建表的表名：",sheet.title

for i in range(1, 10):
    for j in range(1, i + 1):
        sheet.cell(row=i, column=j).value=("%s*%s=%s" % (str(i), str(j), str(i * j)))

all_row=sheet.rows
for i in all_row:
    for j in xrange(len(i)):
        print str(i[j].value).center(10),
    print " "

out.save("D:\\python\\99.xlsx")


