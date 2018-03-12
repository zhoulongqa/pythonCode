#encoding=utf-8
import datetime
import openpyxl
import time
import os
import sys
reload(sys)
sys.setdefaultencoding("utf-8")

workbook=None
excelFile=None
def setWorkBook(excelPathAndFileName):
    global workbook,excelFile
    workbook=openpyxl.load_workbook(excelPathAndFileName)
    excelFile=excelPathAndFileName
    return workbook

#更改表名获取工作表
def getSheetByName(sheetName):
    global workbook
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet
#根据表索引获取工作表
def getSheetByIndex(sheetIndex):
    global workbook
    sheetname=workbook.get_sheet_names()[sheetIndex]
    print "sheet name:",sheetname
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet

#获取sheet中有数据区域的结束行数
def getRowsNumber(sheet):
    return sheet.max_row

#获取sheet中有数据区域的结束列数
def getColsNumber(sheet):
    return sheet.max_column
#获取指定的行
def getRow(sheet,rowNo):
    return sheet.rows[rowNo]
#获取指定单元格的内容
def getCell(sheet,rowNo,colsNo):
    return sheet.cell(row=rowNo,column=colsNo).value
#将内存中的excel保存到硬盘上
def saveExcel():
    global workbook,excelFile
    workbook.save(excelFile)
