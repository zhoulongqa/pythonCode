#encoding=utf-8
from openpyxl import load_workbook
import sys
reload(sys)
#sys.setdefaultencoding("utf-8")
workbook=load_workbook(r"D:\\python\\test.xlsx")
print workbook
#获取excel中所有工作表的名字
bookNameList=workbook.get_sheet_names()
#通过表名获取指定的工作表
sheet1=workbook.get_sheet_by_name(u"员工信息表")

#获取单元格B3
b3=sheet1.cell(coordinate="B3")
print u"修改前单元格B3的值：",b3.value

#修改单元格B3的内容
sheet1.cell(coordinate="B3").value=300
print u"修改后单元格B3的值：",sheet1.cell(coordinate="B3").value

#修改单元格E3的内容
sheet1.cell(row=3,column=5).value="TJ"
print u"修改后单元格B3的值：",sheet1.cell(row=3,column=5).value
workbook.save("D:\\python\\test.xlsx")      #若没有此句则不会修改成功